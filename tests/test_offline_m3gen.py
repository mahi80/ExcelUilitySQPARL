"""Offline test for M3 P4 — ontology / OBDA / SHACL / SPARQL-context generators.
Validates the generated TTL with rdflib. Run: python tests/test_offline_m3gen.py
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import openpyxl
from rdflib import Graph, OWL, RDF, URIRef

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "onboarding"))

from ontology_gen import (  # noqa: E402
    generate_mapping_blocks, generate_ontology_ttl, generate_shapes_ttl,
    generate_sparql_context, project_namespaces)
from profiler import profile_workbook  # noqa: E402

_passed = _failed = 0
TMP = Path(tempfile.mkdtemp())


def check(name, cond, detail=""):
    global _passed, _failed
    if cond:
        _passed += 1; print(f"  [PASS] {name}")
    else:
        _failed += 1; print(f"  [FAIL] {name} — {detail}")


def _local(uri) -> str:
    return str(uri).split("#")[-1]


def main():
    p = TMP / "shop.xlsx"
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    a = wb.create_sheet("authors"); a.append(["author_id", "name"]); a.append(["A1", "Le Guin"])
    b = wb.create_sheet("books")
    b.append(["book_id", "title", "author_id", "price", "in_stock", "published"])
    b.append(["B1", "Earthsea", "A1", 12.5, True, "1968-01-01"])
    b.append(["B2", "1Q84", "A1", 19.0, False, "2009-05-29"])
    wb.save(p)

    prof = profile_workbook(p)
    pid = "p_test01"
    ont_ns, res_ns, ont_prefix, res_prefix = project_namespaces(pid)

    print("ontology.ttl")
    ttl = generate_ontology_ttl(prof, ont_ns, ont_prefix, "shop")
    g = Graph(); g.parse(data=ttl, format="turtle")
    classes = {_local(s) for s in g.subjects(RDF.type, OWL.Class)}
    dprops = {_local(s) for s in g.subjects(RDF.type, OWL.DatatypeProperty)}
    oprops = {_local(s) for s in g.subjects(RDF.type, OWL.ObjectProperty)}
    check("parses + class per table", classes == {"Authors", "Books"}, classes)
    check("datatype properties for all columns",
          {"authorId", "name", "bookId", "title", "price", "inStock", "published"} <= dprops, dprops)
    check("FK -> object property atAuthors", oprops == {"atAuthors"}, oprops)
    price_range = g.value(URIRef(ont_ns + "price"),
                          URIRef("http://www.w3.org/2000/01/rdf-schema#range"))
    check("price range xsd:decimal", str(price_range).endswith("decimal"), price_range)

    print("\nshapes.ttl")
    sg = Graph(); sg.parse(data=generate_shapes_ttl(prof, ont_ns, ont_prefix), format="turtle")
    SH = "http://www.w3.org/ns/shacl#"
    shapes = list(sg.subjects(RDF.type, URIRef(SH + "NodeShape")))
    check("a NodeShape per class", len(shapes) == 2, len(shapes))
    check("shapes reference sh:class for FK",
          (None, URIRef(SH + "class"), URIRef(ont_ns + "Authors")) in sg)

    print("\nmapping.obda blocks")
    blocks = generate_mapping_blocks(prof, f"ds_{pid}", ont_prefix, res_prefix)
    check("one block per table", len(blocks) == 2, len(blocks))
    books_block = next(bk for bk in blocks if "-books" in bk)
    check("source is schema-qualified", f'FROM "ds_{pid}"."books"' in books_block, books_block)
    check("subject uses resource prefix", f"{res_prefix}:books/" in books_block)
    check("class uses ontology prefix", f"{ont_prefix}:Books" in books_block)
    check("FK object triple present", f"{ont_prefix}:atAuthors {res_prefix}:authors/" in books_block)
    check("typed decimal literal", "^^xsd:decimal" in books_block)

    print("\nsparql_context")
    ctx = generate_sparql_context(prof, ont_ns, res_ns, ont_prefix, res_prefix, "shop")
    check("declares project prefix", f"PREFIX {ont_prefix}:" in ctx)
    check("contains PName slash warning", "CANNOT contain '/'" in ctx)
    check("lists classes", f"{ont_prefix}:Books" in ctx)

    print(f"\n{_passed} passed, {_failed} failed")
    return 1 if _failed else 0


if __name__ == "__main__":
    sys.exit(main())
